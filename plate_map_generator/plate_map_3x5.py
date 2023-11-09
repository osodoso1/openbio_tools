import openpyxl
import itertools

# ask the user for the experiment number
exp_num = input("Enter the experiment number: ")
exp_str = "IV0" + exp_num

# ask the user for the excel file path
file_path = input("Enter the excel file path: ")

# load the excel file and select the "Request" tab
workbook = openpyxl.load_workbook(filename=file_path)
worksheet = workbook["Request"]

# extract strain information from the "Request" tab
strains = []
assay_n_sources = []
for row in worksheet.iter_rows(min_row=8, min_col=0, max_col=8):
    strain = row[0].value
    assay_n_source = row[6].value

    if strain is not None:
        strains.append(strain.strip())
    else:
        strains.append(None)

    if assay_n_source == "None":
        assay_n_sources.append("woN")
    else:
        assay_n_sources.append("wN")

# prompt the user for the number of replicates
num_replicates = int(input("Enter the number of replicates (maximum of 5): "))
if num_replicates > 5:
    num_replicates = 5

# generate the strain names
names = []
for i, (strain, assay_n_source) in enumerate(zip(strains, assay_n_sources)):
    if strain is not None:
        for j in range(num_replicates):
            name = f"{exp_str}_{strain}_{assay_n_source}_"
            name += chr(ord('A') + j)
            names.append(name)

# define the plate sizes and capacities
plate_sizes = [(3, 5), (10, 5)]
plate_capacities = [15, 50]

# generate the plate maps
plate_maps = []
plate_num = 1
current_well = 1
for plate_size, plate_capacity in zip(plate_sizes, plate_capacities):
    num_wells_per_plate = plate_size[0] * plate_size[1]
    while current_well <= len(names):
        plate_map = []
        for i in range(plate_size[0]):
            row = []
            for j in range(plate_size[1]):
                if current_well <= len(names):
                    row.append(names[current_well-1])
                current_well += 1
            plate_map.append(row)
        plate_maps.append(plate_map)
        if current_well > len(names):
            break
        plate_num += 1

# save the plate maps and strain list to an excel file
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Plate Maps"
for i, plate_map in enumerate(plate_maps):
    worksheet.append([f"Plate {i+1}"])
    for row in plate_map:
        worksheet.append(row)
    worksheet.append([""])

worksheet = workbook.create_sheet("Strain List")
for name in names:
    worksheet.append([name])

workbook.save(filename="plate_map_3x5.xlsx")
print("Plate maps and strain list saved to plate_map_3x5.xlsx")
